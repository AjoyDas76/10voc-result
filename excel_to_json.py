"""
excel_to_json.py
-----------------
Converts a school mark-entry Excel sheet (subject blocks of
মোট নম্বর / প্রাপ্ত নম্বর / গ্রেড [/ পয়েন্ট বা জি.পি.]) into results.json
for the Result Portal.

Robust to different sheet layouts:
- Works with or without a "Roll" column in the marks sheet itself.
- Reads each subject's actual pass/grade cut-offs directly from that
  column's Excel IF-formula (so it matches the sheet exactly, even if
  different exams/classes use different grading scales).
- If the marks sheet has no Roll column, Roll numbers can be pulled from
  a separate roster file/sheet (name + roll list), matched by row order.

Usage:
    python3 excel_to_json.py <marks.xlsx> [output.json]
        [--sheet SHEET_NAME]
        [--roll-file ROSTER.xlsx --roll-sheet SHEET_NAME]
        [--exam-name "..."] [--exam-class "..."] [--section "..."]
        [--school-name "..."] [--school-address "..."]

Re-run this any time after filling in a new exam's marks to regenerate
results.json. No code changes needed.
"""
import sys
import re
import json
import argparse
import openpyxl

STOP_WORDS = ("জি.পি.এ", "সর্বমোট", "failed", "gpa")

# fallback boundaries (used only if a subject's grade cell has no formula
# to read cut-offs from)
FALLBACK_BOUNDARIES = {
    30: [(23, "A+"), (21, "A"), (17, "A-"), (14, "B"), (11, "C"), (8.5, "D")],
    25: [(17, "A+"), (14.5, "A"), (12.5, "A-"), (10.5, "B"), (8.5, "C"), (6.5, "D")],
    15: [(11, "A+"), (10, "A"), (8, "A-"), (7, "B"), (6, "C"), (4, "D")],
}

FORMULA_RULE_RE = re.compile(r'[A-Z]+\d+\s*(>=|<=|>|<)\s*([\d.]+)\s*,\s*"([^"]*)"')


def parse_grade_rules(ws, sample_row, grade_col):
    """Extract (operator, threshold, label) rules from a grade cell's IF formula."""
    if not grade_col:
        return None
    formula = ws.cell(row=sample_row, column=grade_col).value
    if not isinstance(formula, str) or not formula.startswith("="):
        return None
    rules = [(op, float(val), label) for op, val, label in FORMULA_RULE_RE.findall(formula)]
    return rules or None


def eval_grade(rules, obtained, total):
    if obtained is None:
        obtained = 0
    if rules:
        for op, threshold, label in rules:
            if op == ">=" and obtained >= threshold: return label or "F"
            if op == ">" and obtained > threshold: return label or "F"
            if op == "<=" and obtained <= threshold: return label or "F"
            if op == "<" and obtained < threshold: return label or "F"
        return "F"
    # fallback table
    table = FALLBACK_BOUNDARIES.get(total) or [(t / 30 * total, g) for t, g in FALLBACK_BOUNDARIES[30]]
    for cutoff, letter in table:
        if obtained > cutoff:
            return letter
    return "F"


def find_best_sheet(wb, forced_name=None):
    if forced_name:
        return wb[forced_name]
    best, best_score = None, -1
    for sn in wb.sheetnames:
        ws = wb[sn]
        score = 0
        for r in range(1, min(ws.max_row, 6) + 1):
            for c in range(1, min(ws.max_column, 5) + 1):
                if str(ws.cell(row=r, column=c).value or "").strip() == "নাম":
                    score = ws.max_column  # wider name-bearing sheets win
        if score > best_score:
            best, best_score = ws, score
    if best is None:
        raise ValueError("Could not find a sheet with a 'নাম' header row.")
    return best


def find_layout(ws):
    header_row = sub_row = None
    name_col = roll_col = None
    for r in range(1, min(ws.max_row, 8) + 1):
        for c in range(1, min(ws.max_column, 6) + 1):
            v = str(ws.cell(row=r, column=c).value or "").strip()
            if v == "নাম":
                header_row, name_col = r, c
            elif v.lower() in ("roll", "রোল"):
                roll_col = c
    if header_row is None:
        raise ValueError("Could not find a 'নাম' header cell.")
    sub_row = header_row + 1

    subject_starts = []
    for c in range(name_col + 1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        text = str(v).strip()
        if any(sw in text.lower() for sw in STOP_WORDS):
            break
        subject_starts.append((c, text))

    subjects = []
    for i, (start_col, subj_name) in enumerate(subject_starts):
        end_col = subject_starts[i + 1][0] if i + 1 < len(subject_starts) else ws.max_column + 1
        total_col = obtained_col = grade_col = None
        for c in range(start_col, end_col):
            label = ws.cell(row=sub_row, column=c).value
            if label is None:
                continue
            label = str(label).strip()
            if "মোট" in label:
                total_col = c
            elif "প্রাপ্ত" in label:
                obtained_col = c
            elif "গ্রেড" in label:
                grade_col = c
        if total_col and obtained_col:
            subjects.append((subj_name.strip(), total_col, obtained_col, grade_col))

    return roll_col, name_col, header_row + 2, subjects


def load_roll_map(path, sheet_name):
    """Return an ordered list of (roll, name) from a roster file."""
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    pairs = []
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a is not None and b is not None and str(b).strip():
            pairs.append((str(a).strip(), str(b).strip()))
    return pairs


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("output", nargs="?", default="results.json")
    ap.add_argument("--sheet")
    ap.add_argument("--roll-file")
    ap.add_argument("--roll-sheet")
    ap.add_argument("--exam-name", default=None)
    ap.add_argument("--exam-class", default=None)
    ap.add_argument("--section", default=None)
    ap.add_argument("--exam-id", required=True, help="stable id for this exam, e.g. first_assessment")
    ap.add_argument("--exam-label", default=None, help="text shown in the dropdown, defaults to --exam-name")
    ap.add_argument("--school-name", default="প্রবর্তক স্কুল এন্ড কলেজ")
    ap.add_argument("--school-address", default="পাঁচলাইশ, চট্টগ্রাম")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=False)
    ws = find_best_sheet(wb, args.sheet)

    title_cell = ""
    for c in range(1, 4):
        v = ws.cell(row=1, column=c).value
        if v:
            title_cell = str(v)
            break
    exam_name = args.exam_name or title_cell.split("ফলাফল")[0].strip()
    m_class = re.search(r"শ্রেণিঃ\s*([^\s]+)", title_cell)
    m_sec = re.search(r"শাখাঃ\s*(.+)", title_cell)
    exam_class = args.exam_class or (m_class.group(1).strip() if m_class else "")
    section = args.section or (re.sub(r"\s+", " ", m_sec.group(1).strip()) if m_sec else "")

    roll_col, name_col, first_data_row, subjects = find_layout(ws)
    if not subjects:
        print("Could not detect subject headers — check the sheet layout.")
        sys.exit(1)

    # pre-parse grade formula rules per subject (from the first data row)
    subj_rules = {}
    for subj_name, total_col, obtained_col, grade_col in subjects:
        subj_rules[subj_name] = parse_grade_rules(ws, first_data_row, grade_col)

    roll_lookup = None
    if not roll_col and args.roll_file:
        roll_lookup = load_roll_map(args.roll_file, args.roll_sheet)

    students = []
    auto_roll = 1
    roll_idx = 0
    for row in range(first_data_row, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col).value
        if name is None or str(name).strip() == "":
            continue
        name = str(name).strip()

        if roll_col:
            roll_val = ws.cell(row=row, column=roll_col).value
            roll = str(roll_val).strip() if roll_val not in (None, "") else str(auto_roll)
        elif roll_lookup and roll_idx < len(roll_lookup):
            r_roll, r_name = roll_lookup[roll_idx]
            if r_name != name:
                print(f"WARNING: row {row} name '{name}' != roster name '{r_name}' at position {roll_idx}; using roster roll anyway.")
            roll = r_roll
        else:
            roll = str(auto_roll)
        roll_idx += 1

        subj_rows = []
        for subj_name, total_col, obtained_col, grade_col in subjects:
            total = ws.cell(row=row, column=total_col).value
            obtained = ws.cell(row=row, column=obtained_col).value
            total = total if isinstance(total, (int, float)) else 0
            obtained = obtained if isinstance(obtained, (int, float)) else None
            grade = eval_grade(subj_rules.get(subj_name), obtained, total)
            subj_rows.append({
                "name": subj_name,
                "total": total,
                "obtained": obtained if obtained is not None else 0,
                "grade": grade
            })

        students.append({
            "roll": roll,
            "name": name,
            "class": exam_class,
            "section": section,
            "subjects": subj_rows
        })
        auto_roll += 1

    exam_entry = {
        "id": args.exam_id,
        "label": args.exam_label or exam_name or args.exam_id,
        "examClass": f"{exam_class} শ্রেণি" + (f" ({section})" if section else ""),
        "students": students
    }

    # load existing multi-exam file if present, so we only touch this one exam
    existing = None
    try:
        with open(args.output, "r", encoding="utf-8") as f:
            existing = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        existing = None

    if existing and "exams" in existing:
        data = existing
        data["school"]["name"] = args.school_name or data["school"].get("name")
        data["school"]["address"] = args.school_address or data["school"].get("address")
        exams = data["exams"]
        idx = next((i for i, e in enumerate(exams) if e["id"] == args.exam_id), None)
        if idx is not None:
            exams[idx] = exam_entry
        else:
            exams.append(exam_entry)
    else:
        data = {
            "school": {
                "name": args.school_name,
                "address": args.school_address,
                "logo": "logo.png"
            },
            "exams": [exam_entry]
        }

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(students)} students -> exam '{args.exam_id}' in {args.output}")
    print(f"Roll source: {'sheet column' if roll_col else ('roster file' if roll_lookup else 'auto-numbered')}")
    print(f"Subjects detected: {[s[0] for s in subjects]}")


if __name__ == "__main__":
    main()
